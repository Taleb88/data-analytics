# let's import the following packages
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles.alignment import Alignment

wb = Workbook()
#we will create a new workbook called "employee_list"
excel_file = 'company_data.xlsx'

sheet1 = wb.active #let's make sure we the workbook is active
sheet1.title = "employee_directory" #create a sheet called "employee_list"
sheet2 = wb.create_sheet(title="executive_directory") #create a sheet called "executive_directory"
sheet3 = wb.create_sheet(title="investor_board_directory") #create a sheet called "investor_board_directory"
sheet4 = wb.create_sheet(title="NY_temp_emp_directory") #create a sheet called "NY_temp_emp_directory"
sheet5 = wb.create_sheet(title="LA_temp_emp_directory") #create a sheet called "LA_temp_emp_directory"
sheet6 = wb.create_sheet(title="CHI_temp_emp_directory") #create a sheet called "CHI_temp_emp_directory"
sheet7 = wb.create_sheet(title="MIA_temp_emp_directory") #create a sheet called "MIA_temp_emp_directory"
sheet8 = wb.create_sheet(title="PARIS_temp_emp_directory") #create a sheet called "PARIS_temp_emp_directory"
sheet9 = wb.create_sheet(title="ROME_temp_emp_directory") #create a sheet called "ROME_temp_emp_directory"
sheet10 = wb.create_sheet(title="OSAKA_temp_emp_directory") #create a sheet called "OSAKA_temp_emp_directory"
sheet11 = wb.create_sheet(title="SYDNEY_temp_emp_directory") #create a sheet called "SYDNEY_temp_emp_directory"
sheet12 = wb.create_sheet(title="ROSEHILL_temp_emp_directory") #create a sheet called "ROSEHILL_temp_emp_directory"
sheet13 = wb.create_sheet(title="finances_2020_to_the_present") #create a sheet called "finances_2020_to_the_present"
sheet14 = wb.create_sheet(title="finances_2010_to_2019") #create a sheet called "finances_2010_to_2019"

# data bars added to last column of each worksheet, color = light green
rule = DataBarRule(start_type="percentile",
                   start_value=10,
                   end_type="percentile",
                   end_value=90,
                   color="0000FF00")
# values in a financial column (salary/equity) should now have a light green data bar background in each cell
sheet1.conditional_formatting.add("E2:E100", rule)
sheet2.conditional_formatting.add("E2:E100", rule)
sheet3.conditional_formatting.add("E2:E100", rule)
sheet4.conditional_formatting.add("E2:E100", rule)
sheet5.conditional_formatting.add("E2:E100", rule)
sheet6.conditional_formatting.add("E2:E100", rule)
sheet7.conditional_formatting.add("E2:E100", rule)
sheet8.conditional_formatting.add("E2:E100", rule)
sheet9.conditional_formatting.add("E2:E100", rule)
sheet10.conditional_formatting.add("E2:E100", rule)
sheet11.conditional_formatting.add("E2:E100", rule)
sheet12.conditional_formatting.add("E2:E100", rule)

#columns - name, title, phone, email, salary
#[1:7] = values
employee_data = [
    ["Name","Title","Phone Ext.","Email","Salary *USD*"],
    ["Jaime Rosales","Senior Software Architect",9999, "jrosales@company.com", 155000],
    ["Zach Jamison", "Software Developer", 2828,"zjamison@company.com", 135000],
    ["Danny Kramer", "Cameraman/Photographer", 6996, "dkramer@company.com", 120000],
    ["Muhammad Ali", "Boxing Expert", 1717, "mali@company.com", 71000],
    ["Bill Gates", "Software Expert", 3232, "bgates@company.com", 64000],
    ["Dwayne 'The Rock' Johnson", "Actor", 7272, "djohnson@company.com", 55000],
    ["Lisa Leslie", "WNBA Statistician", 7832, "lleslie@company.com", 65000],
    ["Jennifer Lopez", "Actress", 2469, "jlopez@company.com", 55000],
    ["Kaity Tong", "Journalist", 4747, "ktong@company.com", 61000],
    ["Daniel Day", "Wardrobe Expert", 8844, "dday@company.com", 100000],
    ["Jose Fernandez", "Subject Matter Expert", 7312, "jfernandez@company.com", 55000]
]

executive_data = [
    ["Name","Title","Phone Ext.","Email","Salary *USD*"],
    ["Taleb Damaree", "Chief Executive Officer", 1, "tdamaree@company.com", 750000],
    ["Richard Pryor", "Chief Comedian Officer", 2, "rpryor@company.com", 550000],
    ["Beverly Johnson", "Chief Marketing Officer", 3, "bjohnson@company.com", 510000],
    ["Martin Luther King Jr., PhD", "Chief Civil Rights Officer", 4, "mking@company.com", 480000],
    ["Patrick Ewing", "Chief Athletics Officer", 4, "pewing@company.com", 480000],
    ["Lisa Salters", "Chief People Officer", 5, "lsalters@company.com", 475000],
    ["Harvey Milk", "Chief Equal Opportunity Officer", 6, "hmilk@company.com", 450000],
    ["Junior Bridgeman", "President - Operations", 7, "jbridgeman@company.com", 310000],
    ["Hideo Nomo", "President - Project Management", 8, "hnomo@company.com", 310000],
    ["Martin Lawrence", "President - Marketing", 9, "mlawrence@company.com", 310000],
    ["Diego Maradona", "President - Data Analytics", 10, "dmaradona@company.com", 310000],
    ["Michael Jordan", "President - Sports", 11, "mjordan@company.com", 310000],
    ["Jordan Peele", "President - Entertainment", 12, "jpeele@company.com", 310000],
    ["Al Sharpton", "President - Civil Rights", 13, "sharpton@company.com", 310000],
    ["Paulie Walnuts", "Vice President - Operations", 14, "pwalnuts@company.com", 275000],
    ["Nelson Muntz", "Vice President - Project Management", 15, "nmuntz@company.com", 275000],
    ["George Lopez", "Vice President - Marketing", 16, "glopez@company.com", 275000],
    ["Ross Bagdasarian Jr.", "Vice President - Data Analytics", 17, "rbagdasarian@company.com", 275000],
    ["Todd Smith", "Vice President - Sports", 18, "tsmith@company.com", 275000],
    ["Zack Morris", "Vice President - Entertainment", 19, "zmorris@company.com", 275000],
    ["Coretta Scott King", "Vice President - Civil Rights", 20, "cscottking@company.com", 275000]
]

investor_board_data = [
    ["Name","Occupation","Phone Number","Email","Equity (%) *USD*", "Place of Residence"],
    ["Mark Cuban","Entrepreneur","111-111-1111","mcuban@company.com", 3.5, "Dallas, TX - United States"],
    ["Optimus Prime", "Transformer", "222-222-2222", "oprime@company.com", 3.4, "Dallas, TX - United States"],
    ["Kevin Hart", "Comedian", "333-333-3333", "khart@company.com", 3.4, "Dallas, TX - United States"],
    ["Sadaharu Oh", "Retired Professional Baseball player", "444-444-4444", "soh@company.com", 3.2, "Dallas, TX - United States"],
    ["Ice-T", "Actor/Rapper/Entertainer", "555-555-5555", "icet@company.com", 2.8, "Dallas, TX - United States"],
    ["Daymond John", "Entrepreneur", "666-666-6666", "djohn@company.com", 1.7, "Dallas, TX - United States"]
]

new_york_temp_employee_data = [
    ["Name","Title","Phone Ext.","Email","Salary (Hourly) *USD*"],
    ["Consuela Martinez", "5th Floor Receptionist", 111, "cmartinez@company.com", 28.32],
    ["Frank Thomas", "Custodian", 112, "fthomas@company.com", 27.15],
    ["Hideki Ozawa", "Elevator Operator", 113, "hozawa@company.com", 51.29],
    ["Paul Drake", "Carpenter", 114, "pdrake@company.com", 34.57],
    ["Steve Carrell", "Financial Analyst", 115, "scarrell@company.com", 58.13],
    ["Olena Sabiev", "Software Engineer", 116, "osabiev@company.com", 67.18]
]

los_angeles_temp_employee_data = [
    ["Name","Title","Phone Ext.","Email","Salary (Hourly) *USD*"],
    ["Jose Ferrer", "2nd Floor Receptionist", 222, "jferrer@company.com", 26.11],
    ["Ahmed Bridgewater", "Maintainence", 223, "aatwater@company.com", 31.77],
    ["Rosdam Bibian", "Junior Web Developer", 224, "rbibian@company.com", 47.56],
    ["Ronalda Dos Equis", "Solutions Architect", 225, "rdosequis@company.com", 67.21],
    ["Ryann Parker", "Administrative Assistant", 226, "rparker@company.com", 18.99],
    ["Doron Mordecai", "Administrative Assistant", 227, "dmordecai@company.com", 18.99]
]

chicago_temp_employee_data = [
    ["Name","Title","Phone Ext.","Email","Salary (Hourly) *USD*"],
    ["Rachel Smith", "3rd Floor Receptionist", 331, "rsmith@company.com", 28.67],
    ["Lydia Mtume", "Executive Assistant", 332, "lmtume@company.com", 36.32],
    ["Marian Ibrahimovic", "Senior Mobile App Developer", 333, "mibrahimovic@company.com", 48.79],
    ["Gregory Hayes", "Data Analyst", 334, "ghayes@company.com", 41.44],
    ["Bo Derek", "Administrative Assistant", 335, "bderek@company.com", 20.05],
    ["Don Mattingly", "Senior Administrative Assistant", 336, "dmattingly@company.com", 22.88]
]

miami_temp_employee_data = [
    ["Name","Title","Phone Ext.","Email","Salary (Hourly) *USD*"],
    ["Bailey Hasselhoff", "8th Floor Receptionist", 88881, "bhasselhoff@company.com", 15.43],
    ["Rondell Sheridan", "Executive Administrative Assistant", 88882, "rsheridan@company.com", 25.12],
    ["Mickey Badoo", "Administrative Assistant", 88883, "mbadoo@company.com", 15.43],
    ["Anwar Kapur", "Administrative Assistant", 88884, "akapur@company.com", 15.43],
    ["Emma Stone", "Senior Administrative Assistant", 88885, "estone@company.com", 17.66],
    ["Cheng Xi", "Administrative Assistant", 88886, "cxi@company.com", 15.43]
]

paris_france_temp_employee_data = [
    ["Name","Title","Phone Ext.","Email","Salary (Hourly) *EURO*"],
    ["Emile Abboud", "1st Floor Receptionist", "P111", "eabboud@company.com", 13.10],
    ["Maureen Toussaint", "Executive Assistant", "P112", "mtoussaint@company.com", 22.14],
    ["Wang ZhiZhi", "Data Analyst Manager", "P113", "wzhizhi@company.com", 29.76],
    ["Genevieve Saint-Louis", "Administrative Assistant", "P114", "gsaintlouis@company.com", 13.10],
    ["Kareem Sharif", "Administrative Assistant", "P115", "ksharif@company.com", 13.10],
    ["Lucia Benoit", "Administrative Assistant", "P116", "lbenoit@company.com", 13.10]
]

rome_italy_temp_employee_data = [
    ["Name","Title","Phone Ext.","Email","Salary (Hourly) *EURO*"],
    ["Maria Alongi", "1st Floor Receptionist", "R111", "malongi@company.com", 13.05],
    ["Marco Polo", "Photographer", "R112", "mpolo@company.com", 15.17],
    ["Paulo Vicardi", "Cleaner", "R113", "pvicardi@company.com", 11.32],
    ["Nathaniel Bicardi", "Administrative Assistant", "R114", "nbicardi@company.com", 13.05],
    ["Helena Santos", "Senior Administrative Assistant", "R115", "nsantos@company.com", 16.16],
    ["Constantino Martinez", "Administrative Assistant", "R116", "cmartinez@company.com", 13.10]
]

osaka_japan_temp_employee_data = [
    ["Name","Title","Phone Ext.","Email","Salary (Hourly) *YEN*"],
    ["Yoko Ono", "8th Floor Receptionist", "O811", "yono@company.com", 902.00],
    ["Rui Nakaido", "Photographer/Cameraman", "O812", "rnakaido@company.com", 1451.98],
    ["Ken Watanabe", "Marketing Specialist", "O813", "kwatanabe@company.com", 978.76],
    ["Hayao Miyazaki", "Administrative Assistant", "O814", "hmiyazaki@company.com", 902.00],
    ["Shinji Kagawa", "Administrative Assistant", "O815", "kagawa@company.com", 902.00],
    ["Hideki Matsui", "Senior Baseball Scout", "O816", "hmatsui@company.com", 2021.34]
]

sydney_australia_temp_employee_data = [
    ["Name","Title","Phone Ext.","Email","Salary (Hourly) *AUD*"],
    ["Paul Hogan", "4th Floor Receptionist", "S411", "phogan@company.com", 17.70],
    ["Gordon Thomas", "Fisherman", "S812", "gthomas@company.com", 17.70],
    ["Michelle Walker", "Marketing Specialist", "S813", "mwalker@company.com", 22.96],
    ["Andrew Gaze", "Athletic Director", "S814", "agaze@company.com", 51.71],
    ["Rachelle Lewis", "Bookkeeper", "S815", "rlewis@company.com", 32.64],
    ["David Gulpilil", "Senior Film Editor", "S816", "dgulpilil@company.com", 48.32]
]

rosehill_mauritius_temp_employee_data = [
    ["Name","Title","Phone Ext.","Email","Salary (Hourly) *Rs*"],
    ["Ayesha Naughtoo", "5th Floor Receptionist", "M511", "anaughtoo@company.com", 22.14],
    ["Rabea Damaree", "Human Resources Consultant", "M512", "rdamaree@company.com", 36.88],
    ["Fareed Ahmed", "IT Consultant", "M512", "fahmed@company.com", 31.71],
    ["Bruno Julie", "Boxing Expert", "M513", "bjulie@company.com", 17.96],
    ["Seewoosagur Ramgoolam", "Administrative Assistant", "M514", "sramgoolam@company.com", 17.10],
    ["Paula Tung", "Accountant", "M515", "ptung@company.com", 25.66],
    ["Nadeem Wachil", "Cameraman", "M516", "nwachil@company.com", 21.43],
    ["Steven Obeegadoo", "Bookkeeper", "M517", "sobeegadoo@company.com", 27.90],
    ["Cai Yuan", "Accountant", "M518", "cyuan@company.com", 25.66]
]

quarterly_financials_2020_to_the_present_data = [
    ["Year", "Quarter 1", "Quarter 2", "Quarter 3", "Quarter 4", "Total"],
    [2020, 11232312.35, 8432918.88, 9744531.69, 15290114.41, "=SUM(B2:E2)"],
    [2021, 18235537.76, 12436924.99, 16722910.11, 220452198.58, "=SUM(B3:E3)"],
    [2022, 24111925.55, 10898542.36, 24778445.92, "TBD", "=SUM(B4:D4)"],
    [" ", " ", " "," ", " ", "=SUM(F2:F4)"] # decade total value
]

quarterly_financials_2010_to_2019 = [
    ["Year", "Quarter 1", "Quarter 2", "Quarter 3", "Quarter 4", "Total"],
    [2010, 4632739.22, 7320381.54, 9038923.71, 13826184.46, "=SUM(B2:E2)"],
    [2011, 18235537.13, 12432924.99, 12722910.15, 220452198.58, "=SUM(B3:E3)"],
    [2012, 24111925.44, 10898542.36, 7743821.03, 17638291.67, "=SUM(B4:D4)"],
    [2013, 11232312.75, 87762342.88, 84723213.84, 15290154.41, "=SUM(B5:E5)"],
    [2014, 18235537.77, 12436928.92, 1717183.11, 220852798.58, "=SUM(B6:E6)"],
    [2015, 24111925.89, 1177239.19, 24743844.90, 203291331.08, "=SUM(B7:D7)"],
    [2016, 11232312.83, 9330213.33, 9744531.69, 332031941, "=SUM(B8:E8)"],
    [2017, 18235537.73, 12436924.99, 16722910.15, 20981342.58, "=SUM(B9:E9)"],
    [2018, 24111925.42, 14042201.59, 30029321.92, 8372931.99, "=SUM(B10:D10)"],
    [2019, 24111925.10, 12058438.81, 98482394.49, 7493232.77, "=SUM(B11:D11)"],
    [" ", " ", " "," ", " ", "=SUM(F2:F11)"] #decade total value
]

''' reading data from a specific cell in worksheet
c = sheet1["B10"]
print(c) # <cell 'employee_directory'.B10>
'''

# let's add/append all data to the sheets with a for loop
for x in employee_data:
    sheet1.append(x)
for x in executive_data:
    sheet2.append(x)
for x in investor_board_data:
    sheet3.append(x)
for x in new_york_temp_employee_data:
    sheet4.append(x)
for x in los_angeles_temp_employee_data:
    sheet5.append(x)
for x in chicago_temp_employee_data:
    sheet6.append(x)
for x in miami_temp_employee_data:
    sheet7.append(x)
for x in paris_france_temp_employee_data:
    sheet8.append(x)
for x in rome_italy_temp_employee_data:
    sheet9.append(x)
for x in osaka_japan_temp_employee_data:
    sheet10.append(x)
for x in sydney_australia_temp_employee_data:
    sheet11.append(x)
for x in rosehill_mauritius_temp_employee_data:
    sheet12.append(x)
for x in quarterly_financials_2020_to_the_present_data:
    sheet13.append(x)
for x in quarterly_financials_2010_to_2019:
    sheet14.append(x)
# bold font style for the headers in all sheets
for cell in sheet1["1:1"]:
    cell.font = Font(bold=True,size=12)
for cell in sheet2["1:1"]:
    cell.font = Font(bold=True,size=12)
for cell in sheet3["1:1"]:
    cell.font = Font(bold=True,size=12)
for cell in sheet4["1:1"]:
    cell.font = Font(bold=True,size=12)
for cell in sheet5["1:1"]:
    cell.font = Font(bold=True,size=12)
for cell in sheet6["1:1"]:
    cell.font = Font(bold=True,size=12)
for cell in sheet7["1:1"]:
    cell.font = Font(bold=True,size=12)
for cell in sheet8["1:1"]:
    cell.font = Font(bold=True,size=12)
for cell in sheet9["1:1"]:
    cell.font = Font(bold=True,size=12)
for cell in sheet10["1:1"]:
    cell.font = Font(bold=True,size=12)
for cell in sheet11["1:1"]:
    cell.font = Font(bold=True,size=12)
for cell in sheet12["1:1"]:
    cell.font = Font(bold=True, size=12)
for cell in sheet13["1:1"]:
    cell.font = Font(bold=True, size=12)
for cell in sheet14["1:1"]:
    cell.font = Font(bold=True, size=12)
# let's align the text in cell E4 of sheet13 to the center
sheet13["E4"].alignment = Alignment(horizontal="center")
# let's bold the end total value in sheet 13 and sheet 14
sheet13["F5"].font = Font(bold=True)
sheet14["F12"].font = Font(bold=True)
# for salary/equity in all sheets, let's make the color dark purple and the style=italic
for cell in sheet1['E']:
    cell.font = Font(bold=True,color="00660066",italic=True)
for cell in sheet2['E']:
    cell.font = Font(bold=True,color="00660066",italic=True)
for cell in sheet3['E']:
    cell.font = Font(bold=True,color="00660066",italic=True)
for cell in sheet4['E']:
    cell.font = Font(bold=True,color="00660066",italic=True)
for cell in sheet5['E']:
    cell.font = Font(bold=True,color="00660066",italic=True)
for cell in sheet6['E']:
    cell.font = Font(bold=True,color="00660066",italic=True)
for cell in sheet7['E']:
    cell.font = Font(bold=True,color="00660066",italic=True)
for cell in sheet8['E']:
    cell.font = Font(bold=True,color="00660066",italic=True)
for cell in sheet9['E']:
    cell.font = Font(bold=True,color="00660066",italic=True)
for cell in sheet10['E']:
    cell.font = Font(bold=True,color="00660066",italic=True)
for cell in sheet11['E']:
    cell.font = Font(bold=True,color="00660066",italic=True)
for cell in sheet12['E']:
    cell.font = Font(bold=True,color="00660066",italic=True)
# let's revert the font color back to black, the style should be bold, and the size=12 for Salary
sheet1["E1"].font = Font(bold=True,size=12,color="00000000")
sheet2["E1"].font = Font(bold=True,size=12,color="00000000")
sheet3["E1"].font = Font(bold=True,size=12,color="00000000")
sheet4["E1"].font = Font(bold=True,size=12,color="00000000")
sheet5["E1"].font = Font(bold=True,size=12,color="00000000")
sheet6["E1"].font = Font(bold=True,size=12,color="00000000")
sheet7["E1"].font = Font(bold=True,size=12,color="00000000")
sheet8["E1"].font = Font(bold=True,size=12,color="00000000")
sheet9["E1"].font = Font(bold=True,size=12,color="00000000")
sheet10["E1"].font = Font(bold=True,size=12,color="00000000")
sheet11["E1"].font = Font(bold=True,size=12,color="00000000")
sheet12["E1"].font = Font(bold=True,size=12,color="00000000")
# creating filters for the headers
sheet1.auto_filter.ref = "A1:E3"
sheet2.auto_filter.ref = "A1:E3"
sheet3.auto_filter.ref = "A1:F3"
sheet4.auto_filter.ref = "A1:E3"
sheet5.auto_filter.ref = "A1:E3"
sheet6.auto_filter.ref = "A1:E3"
sheet7.auto_filter.ref = "A1:E3"
sheet8.auto_filter.ref = "A1:E3"
sheet9.auto_filter.ref = "A1:E3"
sheet10.auto_filter.ref = "A1:E3"
sheet11.auto_filter.ref = "A1:E3"
sheet12.auto_filter.ref = "A1:E3"
sheet13.auto_filter.ref = "A1:F1"
sheet14.auto_filter.ref = "A1:F1"
# all for sorting
sheet1.auto_filter.add_sort_condition("A2:E3")
sheet2.auto_filter.add_sort_condition("A2:E3")
sheet3.auto_filter.add_sort_condition("A2:F3")
sheet4.auto_filter.add_sort_condition("A2:E3")
sheet5.auto_filter.add_sort_condition("A2:E3")
sheet6.auto_filter.add_sort_condition("A2:E3")
sheet7.auto_filter.add_sort_condition("A2:E3")
sheet8.auto_filter.add_sort_condition("A2:E3")
sheet9.auto_filter.add_sort_condition("A2:E3")
sheet10.auto_filter.add_sort_condition("A2:E3")
sheet11.auto_filter.add_sort_condition("A2:E3")
sheet12.auto_filter.add_sort_condition("A2:E3")
sheet13.auto_filter.add_sort_condition("A1:F1")
sheet14.auto_filter.add_sort_condition("A1:F1")
# let's save all the data into the workbook
wb.save(filename=excel_file)