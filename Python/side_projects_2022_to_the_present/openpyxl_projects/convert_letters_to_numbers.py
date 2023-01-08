import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
#convert from str to int
print(get_column_letter(1)) #A
print(get_column_letter(5)) #E

#lets open an excel file
wb = openpyxl.load_workbook('file_example_XLS_5000.xlsx')
sheet = wb['Sheet2']
#convert int to str
print(sheet.max_column) #8
print(column_index_from_string('D')) #4
print(column_index_from_string('DD')) #108
