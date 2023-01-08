import openpyxl

wb = openpyxl.load_workbook('file_example_XLS_5000.xlsx')

sheet = wb['Sheet3'] #get Sheet3 from workbook

print(sheet.max_row) #get highest row number -> 701
print(sheet.max_column) #get highest column number -> 16