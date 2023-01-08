import openpyxl

wb = openpyxl.load_workbook('file_example_XLS_5000.xlsx')

sheet = wb['Sheet3'] #get Sheet3 from workbook
sheet['A1'] #get cell A1 from sheet (Sheet 3 in this case)

c = sheet['C18'] #get cell C18 from sheet (sheet 3 in this case)

print(sheet['A3'].value) #get value from cell
print('Row %s, Column %s is %s' %(c.row, c.column, c.value)) #get row, comma, value from the cell

print(sheet.cell(row=4, column=11)) #<Cell 'Sheet3'.K4>
print(sheet.cell(row=4, column=11).value) #21780

for x in range(2, 15, 3): # go through every 3 rows up to row 15
    print(x, sheet.cell(row=x, column=x).value) #provides a list containing column numbers and names

