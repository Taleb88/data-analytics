from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()

dest_filename = 'test.xlsx'

sheet1 = wb.active
sheet1.title = "range names"

for row in range(1, 40):
    sheet1.append(range(1000))

sheet2 = wb.create_sheet(title="test")

sheet2['G8'] = 'Taleb'

sheet3 = wb.create_sheet(title='test2')
for row in range(10, 20):
    for col in range(27, 54):
        _=sheet3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
print(sheet3['BB55'].value)
wb.save(filename=dest_filename)